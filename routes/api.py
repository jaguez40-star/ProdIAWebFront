"""
API routes - Clean version with direct template matching
"""

from flask import Blueprint, request, jsonify, session
import requests
import logging
import time
import os
import json
import sqlite3
import csv
import threading
import unicodedata
from datetime import datetime
from collections import defaultdict
from utils.chat_history import chat_history_manager

# Cache de reportes fijos con TTL (los datos se actualizan ~1 vez/día)
_report_cache = {}  # {cache_key: {"data": dict, "timestamp": float}}
REPORT_CACHE_TTL = 300  # 5 minutos

api_bp = Blueprint("api", __name__)
INGESTA_API_URL = os.environ.get("INGESTA_API_URL", "http://localhost:5030")
logger = logging.getLogger(__name__)


@api_bp.route("/ingesta/upload", methods=["POST"])
def ingesta_upload():
    """Proxy: reenvía el archivo subido al FastAPI de INGESTA para su ingesta en daily_report_prod."""
    if "file" not in request.files:
        return jsonify({"success": False, "error": "no se envió ningún archivo"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"success": False, "error": "archivo sin nombre"}), 400
    try:
        resp = requests.post(
            f"{INGESTA_API_URL}/ingesta/upload",
            files={"file": (f.filename, f.stream, f.mimetype or "application/octet-stream")},
            timeout=900,  # archivos NEW grandes pueden tardar varios minutos
        )
        try:
            payload = resp.json()
        except ValueError:
            payload = {"success": False, "error": f"respuesta no-JSON de INGESTA: {resp.text[:300]}"}
        return jsonify(payload), resp.status_code
    except requests.RequestException as e:
        return jsonify({"success": False, "error": f"INGESTA no disponible en {INGESTA_API_URL}: {e}"}), 502


@api_bp.route("/ingesta/check_existing")
def ingesta_check_existing():
    """Proxy: verifica si ya existe un reporte para la fecha dada."""
    try:
        resp = requests.get(f"{INGESTA_API_URL}/ingesta/check_existing",
                            params={"fecha": request.args.get("fecha", "")}, timeout=10)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/ingesta/upload_stream", methods=["POST"])
def ingesta_upload_stream():
    """Reenvía el archivo al SSE de INGESTA, retransmite el progreso por SocketIO y DEVUELVE el
    resultado final como respuesta HTTP (cierre determinista; el 'fin' no depende de un emit suelto)."""
    from flask import current_app
    if "file" not in request.files:
        return jsonify({"success": False, "error": "no se envió ningún archivo"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"success": False, "error": "archivo sin nombre"}), 400
    user_id = session.get("user_id") or (session.get("user", {}) or {}).get("username", "default")
    socketio = current_app.extensions["socketio"]
    files = {"file": (f.filename, f.read(), f.mimetype or "application/octet-stream")}
    final = {"success": False, "error": "sin respuesta de INGESTA"}
    try:
        with requests.post(f"{INGESTA_API_URL}/ingesta/upload_stream",
                           files=files, stream=True, timeout=1800) as resp:
            if resp.status_code != 200:
                return jsonify({"success": False,
                                "error": f"INGESTA respondió {resp.status_code}: {resp.text[:300]}"}), resp.status_code
            resp.encoding = "utf-8"   # nombres de hoja con tilde correctos (G1 del plan v3)
            for line in resp.iter_lines(decode_unicode=True):
                if not line or not line.startswith("data:"):
                    continue
                try:
                    ev = json.loads(line[5:].strip())
                except ValueError:
                    continue
                if ev.get("tipo") == "fin":
                    final = {"success": ("error" not in ev), **ev}   # capturar, NO emitir
                else:
                    socketio.emit("ingesta_progress", ev, room=user_id)
    except requests.RequestException as e:
        return jsonify({"success": False, "error": f"INGESTA no disponible en {INGESTA_API_URL}: {e}"}), 502
    return jsonify(final), (200 if final.get("success") else 500)


@api_bp.route("/tablas-hoja")
def tablas_hoja_listar():
    """Proxy: lista de tablas lógicas de una hoja (reporte_id, hoja)."""
    try:
        resp = requests.get(f"{INGESTA_API_URL}/tablas",
                            params={"reporte_id": request.args.get("reporte_id"),
                                    "hoja": request.args.get("hoja")}, timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502

@api_bp.route("/tablas-hoja/datos")
def tablas_hoja_datos():
    """Proxy: contenido ancho de una tabla (reporte_id, hoja, tabla_idx)."""
    try:
        resp = requests.get(f"{INGESTA_API_URL}/tablas/datos",
                            params={"reporte_id": request.args.get("reporte_id"),
                                    "hoja": request.args.get("hoja"),
                                    "tabla_idx": request.args.get("tabla_idx")}, timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/tablas-hoja/arbol")
def tablas_hoja_arbol():
    """Proxy: árbol jerárquico de reportes ingeridos (metadata, sin hojas -- ver /arbol/<reporte_id>)."""
    try:
        resp = requests.get(f"{INGESTA_API_URL}/tablas/arbol", timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/tablas-hoja/arbol/<int:reporte_id>")
def tablas_hoja_arbol_detalle(reporte_id):
    """Proxy: desglose de hojas/tablas de un reporte especifico (lazy-load al expandir un dia)."""
    try:
        resp = requests.get(f"{INGESTA_API_URL}/tablas/arbol/{reporte_id}", timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/reportes/ultimo")
def reportes_ultimo():
    """Proxy: fecha del reporte más reciente cargado en la BD (badge 'Act:' del panel Insights)."""
    try:
        resp = requests.get(f"{INGESTA_API_URL}/reportes/ultimo", timeout=15)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/analisis/catalogo")
def analisis_catalogo():
    """Proxy: catálogo de entidades (cardinalidad + colisiones) desde INGESTA."""
    try:
        resp = requests.get(f"{INGESTA_API_URL}/analisis/catalogo", timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


# ---- Caché TTL + single-flight para los paneles de análisis (2026-08-04) --------------------------
# POR QUÉ: /analisis/ejecutivo NO tiene caché en FastAPI (verificado) → cada petición re-invoca a Gemma
# (~180s con EJECUTIVO_USAR_LLM=true). Con la precarga desde el login, CADA login dispararía una
# generación; como Ollama serializa, N logins casi simultáneos se encolan y el último revienta el
# timeout=200 de este proxy. Además app.py usa async_mode="threading" con `requests` BLOQUEANTE → cada
# petición en vuelo ocupa un hilo.
# QUÉ HACE: (1) TTL — el payload global se reutiliza N minutos; (2) single-flight — si M peticiones
# piden la MISMA clave y no hay caché, solo UNA llama al backend y las otras esperan y reciben su
# resultado (evita la estampida). Solo se cachean respuestas BUENAS (200 + sin marca de error).
# NO toca FastAPI: `analisis/api.py::ejecutivo()` queda intacto (decisión AF-A2) y el grupo Analizar,
# que la llama in-process con pulir=False, NO pasa por aquí → sin contaminación cruzada de payloads.
_ANALISIS_TTL_S = int(os.getenv("ANALISIS_CACHE_TTL", "900"))   # 15 min; el reporte cambia 1 vez/día
_ANALISIS_CACHE = {}            # clave -> (expira_ts, payload, status)
_ANALISIS_GUARD = threading.Lock()      # protege _ANALISIS_CACHE y _ANALISIS_INFLIGHT
_ANALISIS_INFLIGHT = {}         # clave -> threading.Lock (uno por clave)


def _analisis_cache_get(clave):
    """Devuelve (payload, status) si hay entrada viva; None si no hay o expiró."""
    with _ANALISIS_GUARD:
        v = _ANALISIS_CACHE.get(clave)
        if not v:
            return None
        if v[0] <= time.time():
            _ANALISIS_CACHE.pop(clave, None)
            return None
        return v[1], v[2]


def _analisis_inflight_lock(clave):
    with _ANALISIS_GUARD:
        lk = _ANALISIS_INFLIGHT.get(clave)
        if lk is None:
            lk = threading.Lock()
            _ANALISIS_INFLIGHT[clave] = lk
        return lk


def _analisis_es_cacheable(payload, status):
    """Solo se cachean respuestas útiles. MISMO criterio que las guardas del frontend: nunca cachear
    un error (dejaría el panel mostrando basura sin reintentar hasta que expire el TTL)."""
    if status != 200 or not isinstance(payload, dict):
        return False
    if payload.get("encontrada") is False or payload.get("sin_datos") or payload.get("error"):
        return False
    if (payload.get("meta") or {}).get("generado_por") == "error":   # Gemma falló (ejecutivo)
        return False
    # president: encontrada=True pero productos=[] es el "Compromiso P50 no disponible" del frontend
    # (__cnPaintP50Header) — no cachear ese header vacío. Solo president trae la clave `productos`;
    # desempeno usa `por_producto` y ejecutivo `gap_por_producto`, así que esto no los afecta.
    if "productos" in payload and not payload.get("productos"):
        return False
    return True


# [2026-08-25] `pulir` NO entra en la clave: solo cambia el texto de `secciones` (prosa de Gemma),
# no las cifras/focos/tarjetas. Si formara parte de la clave, el panel de grano día (pulir=false) y
# el tablero (sin `pulir` -> True) serían DOS entradas para el mismo payload: el panel nunca
# reutilizaría lo que el tablero ya cargó y el single-flight no los agruparía. Esa era la causa de
# que el panel tardara minutos con el tablero ya pintado.
_CLAVE_IGNORAR = ("pulir",)


def _analisis_proxy_cacheado(ruta, params, timeout):
    """Proxy con caché TTL + single-flight. `ruta` es el path en INGESTA (ej. '/analisis/ejecutivo')."""
    clave = ruta + "?" + "&".join(f"{k}={params[k]}" for k in sorted(params)
                                  if k not in _CLAVE_IGNORAR)
    hit = _analisis_cache_get(clave)
    if hit:
        return jsonify(hit[0]), hit[1]
    lock = _analisis_inflight_lock(clave)
    with lock:                                  # single-flight: solo 1 generación por clave
        hit = _analisis_cache_get(clave)        # double-check: otro hilo pudo llenarla mientras esperábamos
        if hit:
            return jsonify(hit[0]), hit[1]
        try:
            resp = requests.get(f"{INGESTA_API_URL}{ruta}", params=params, timeout=timeout)
            payload, status = resp.json(), resp.status_code
        except requests.RequestException as e:
            return jsonify({"error": f"INGESTA no disponible: {e}"}), 502
        # Un payload SIN pulir se sirve, pero no llena la entrada compartida: el tablero pide la
        # misma clave y sí muestra la prosa de `secciones`. Al revés sí vale — el panel de día
        # ignora `secciones`, así que reutiliza sin problema lo que dejó el tablero.
        _sin_pulir = str(params.get("pulir", "")).lower() in ("false", "0")
        if _analisis_es_cacheable(payload, status) and not _sin_pulir:
            with _ANALISIS_GUARD:
                _ANALISIS_CACHE[clave] = (time.time() + _ANALISIS_TTL_S, payload, status)
        return jsonify(payload), status
# ---- fin caché de análisis -----------------------------------------------------------------------


@api_bp.route("/analisis/president")
def analisis_president():
    """Proxy: tarjeta P50 por producto (REPORTE_PRESIDENT, escala kbpe corporativa) desde INGESTA."""
    try:
        params = {}
        per = request.args.get("periodo")
        if per:
            params["periodo"] = per
        return _analisis_proxy_cacheado("/analisis/president", params, 30)
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/analisis/desempeno")
def analisis_desempeno():
    """Proxy: desempeño del mes (KPIs REAL vs PPTO + curva diaria) de una entidad."""
    try:
        params = {}
        ent = request.args.get("entidad")
        if ent:
            params["entidad"] = ent   # H3: idiom del repo — solo pasar entidad si viene (evita ""→no encontrada)
        seg = request.args.get("segmento")
        if seg:
            params["segmento"] = seg
        for _k in ("nivel", "periodo"):
            _v = request.args.get(_k)
            if _v:
                params[_k] = _v
        return _analisis_proxy_cacheado("/analisis/desempeno", params, 45)
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/ebitda/unificado-waterfall")
def ebitda_unificado_waterfall():
    """Proxy: waterfall EBITDA (Ingresos→NOPAT) desde ROBUSTEZ/ops. global | activo | campo."""
    try:
        params = {}
        for _k in ("year", "month", "nivel", "entidad"):
            _v = request.args.get(_k)
            if _v:
                params[_k] = _v
        resp = requests.get(f"{INGESTA_API_URL}/ebitda/unificado-waterfall", params=params, timeout=45)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/analisis/desempeno_insight")
def analisis_desempeno_insight():
    """Proxy: titular ejecutivo IA (contrato JSON) del desempeño del mes."""
    try:
        params = {}
        ent = request.args.get("entidad")
        if ent:
            params["entidad"] = ent
        seg = request.args.get("segmento")
        if seg:
            params["segmento"] = seg
        for _k in ("nivel", "periodo"):
            _v = request.args.get(_k)
            if _v:
                params[_k] = _v
        resp = requests.get(f"{INGESTA_API_URL}/analisis/desempeno_insight", params=params, timeout=90)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/analisis/ejecutivo")
def analisis_ejecutivo():
    """Proxy: Análisis Ejecutivo IA multi-sección (insights/oportunidades/puntos_atencion/decisiones).
    Sandbox tarjeta Filiales (EN PRUEBAS) -- timeout=200 (> 180s del LLM en FastAPI, on-demand)."""
    try:
        params = {}
        ent = request.args.get("entidad")
        if ent:
            params["entidad"] = ent
        seg = request.args.get("segmento")
        if seg:
            params["segmento"] = seg
        # [2026-08-25] `pulir` se reenvía (QV2-PANEL-DIA, A-5): sin él INGESTA usa pulir=True y
        # ejecuta el pulido LLM de Gemma (timeout 200s aquí). El panel de grano día NO usa
        # `secciones`, así que pide pulir=false y responde en el acto. Analizar ya evitaba ese
        # coste llamando en proceso (respuesta_analizar.py:253, RA-1); esta es la vía HTTP.
        for _k in ("nivel", "periodo", "pulir"):
            _v = request.args.get(_k)
            if _v:
                params[_k] = _v
        return _analisis_proxy_cacheado("/analisis/ejecutivo", params, 200)
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/analisis/densidad")
def analisis_densidad():
    """Proxy: auditoría de densidad temporal desde INGESTA. Con ?entidad= → filtra por fuentes ECP."""
    try:
        params = {}
        ent = request.args.get("entidad")
        if ent:
            params["entidad"] = ent
        resp = requests.get(f"{INGESTA_API_URL}/analisis/densidad", params=params, timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/analisis/tendencia_filial")
def analisis_tendencia_filial():
    """Proxy: tendencia de UNA filial (Hocol/America/Permian) — proyección de cierre vs su promedio 2026.
    Panel EXCLUSIVO de la filial (no el panorama de las 3)."""
    try:
        params = {}
        emp = request.args.get("empresa")
        if emp:
            params["empresa"] = emp
        per = request.args.get("periodo")
        if per:
            params["periodo"] = per
        resp = requests.get(f"{INGESTA_API_URL}/analisis/tendencia_filial", params=params, timeout=45)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/analisis/huella")
def analisis_huella():
    """Proxy: huella de datos por fact/escenario (global si no hay entidad, o por entidad)."""
    try:
        params = {}
        ent = request.args.get("entidad")
        if ent:
            params["entidad"] = ent
        resp = requests.get(f"{INGESTA_API_URL}/analisis/huella", params=params, timeout=45)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/analisis/cobertura")
def analisis_cobertura():
    """Proxy: cobertura del reporte (todas las hojas por categoría). Con ?entidad= → presencia por hoja."""
    try:
        params = {}
        ent = request.args.get("entidad")
        if ent:
            params["entidad"] = ent
        resp = requests.get(f"{INGESTA_API_URL}/analisis/cobertura", params=params, timeout=90)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


# ============================================================================
# Diferidas históricas por causa — ruta Flask NATIVA (NO proxy a INGESTA).
# Fuente: data/ECP_DIFERIDAS/ECP_DIFERIDAS.db (SQLite local, ya saneada, read-only).
# Frecuencia por INCIDENTES (colapsa las filas diarias de un mismo evento por INI/END),
# en % dentro de la entidad (campo/activo). El grano de AVM_DATADIF es día-pozo; el año
# fiable es EVENT_DATE (INI/END tienen formato mixto). Ver plan Diferidas.
# ============================================================================
# Raíz del repo ANCLADA con abspath: independiente del CWD y de si __file__ llega relativo
# (la app se lanza como `python app.py`, ruta relativa → sin abspath, dirname(dirname()) colapsaba
# a "" y os.path.exists daba False → "BD no disponible" aunque el archivo sí existe).
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DIF_DB = os.path.join(_BASE_DIR, "data", "ECP_DIFERIDAS", "ECP_DIFERIDAS.db")
_ACTIVO_CSV = os.path.join(_BASE_DIR, "data", "Activo_campo.csv")
# Caché en memoria de /diferidas/frecuencia: los datos son HISTÓRICOS/ESTÁTICOS (ene-2023 → jul-2025),
# así que el resultado por entidad no cambia → se calcula una vez y las reaperturas son instantáneas.
_DIF_CACHE = {}


def _dif_campos_de_activo(activo):
    """Campos que componen un ACTIVO (data/Activo_campo.csv, formato ACTIVO;CAMPO)."""
    out, up = [], (activo or "").strip().upper()
    try:
        with open(_ACTIVO_CSV, encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh, delimiter=";"):
                if (row.get("ACTIVO") or "").strip().upper() == up:
                    c = (row.get("CAMPO") or "").strip()
                    if c:
                        out.append(c)
    except OSError:
        pass
    return out


# ============================================================================
# Mantenimientos (Eventos OW): eventos de servicio a pozo, pill "Mantenimientos" del
# acordeón de foco. Mismo patrón que Diferidas (ruta NATIVA, singleton en memoria,
# degradación SIEMPRE HTTP 200). Ver plan_mantenimientos_eventos_ow_2026-08-13.md.
# ============================================================================
_MTTO_XLSX = os.path.join(_BASE_DIR, "data", "Eventos_OW.xlsx")
_MTTO_ROWS = None                    # singleton lazy: el .xlsx se parsea 1 sola vez por proceso
_MTTO_LOCK = threading.Lock()        # async_mode="threading" + parseo de ~1,5s -> single-flight
_MTTO_MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
               "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _mtto_norm(s):
    """UPPER + trim + plegar acentos/ñ. MISMO criterio que consulta_v2/normaliza.py::norm.
    El .xlsx viene en NFC ('CAÑO SUR ESTE') y nadie normaliza del lado de Postgres -> el match
    literal sería byte a byte y una fuente en NFD rompería campos EN SILENCIO. Verificado que
    hoy da el mismo resultado que el literal (92 campos = 92) -> es blindaje sin cambio de conducta."""
    s = unicodedata.normalize("NFKD", (s or "").strip().upper())
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.split())


def _mtto_cargar():
    """Parsea Eventos_OW.xlsx UNA vez por proceso. None = archivo ausente (se reintenta en la
    siguiente llamada, por si aparece tras un despliegue).

    🔑 `FinalizaEvento` VACÍO significa evento ABIERTO (sin cierre reportado), NO fila inválida
    — son 3.305 de 6.850 filas (48%). Descartarlas perdería justo los eventos que siguen
    corriendo. 🔑 Doble chequeo bajo lock — el parseo tarda ~1,5 s y Flask sirve con hilos."""
    global _MTTO_ROWS
    if _MTTO_ROWS is not None:
        return _MTTO_ROWS
    if not os.path.exists(_MTTO_XLSX):
        return None
    with _MTTO_LOCK:
        if _MTTO_ROWS is not None:            # otro hilo lo cargó mientras esperábamos
            return _MTTO_ROWS
        import openpyxl
        wb = openpyxl.load_workbook(_MTTO_XLSX, data_only=True, read_only=False)
        ws = wb["Sheet1"]
        out = []
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            campo, pozo, ini, fin = r[3], r[4], r[7], r[8]
            if not campo or not isinstance(ini, datetime):
                continue
            # 5 filas traen el año mal tecleado (2526/2626/3026/2016) -> se tratan como abiertas
            f = fin if (isinstance(fin, datetime) and 2020 <= fin.year <= 2030) else None
            tipo = (r[12] or r[10] or r[6] or "").strip() or "Evento"   # Objetivo Ppal → TIPO_EVENTO → Evento
            out.append({"campo": _mtto_norm(campo), "pozo": (pozo or "").strip(),
                        "tipo": tipo, "inicio": ini, "fin": f})
        _MTTO_ROWS = out
        return out


def _mtto_mes(periodo):
    """'2026-05' | 'Mayo 2026' -> (2026, 5). None si no se puede derivar.

    🔑 Valida RANGO, no solo formato: '2026-13' calza el patrón \\d{4}-\\d{2} pero datetime(y,13,1)
    lanza ValueError -> la ruta devolvería HTTP 500 y rompería el contrato "SIEMPRE 200". Un mes
    fuera de 1..12 (o un año absurdo) se trata como periodo NO derivable -> cae al mes más
    reciente con eventos, que es la degradación ya prevista."""
    p = (periodo or "").strip()
    if not p:
        return None
    y = m = None
    if len(p) == 7 and p[4] == "-" and p[:4].isdigit() and p[5:].isdigit():
        y, m = int(p[:4]), int(p[5:])
    else:
        partes = p.split()
        if len(partes) == 2 and partes[1].isdigit():
            nom = partes[0].capitalize()
            if nom in _MTTO_MESES:
                y, m = int(partes[1]), _MTTO_MESES.index(nom)
    if y is None or not (1 <= m <= 12) or not (1900 <= y <= 2999):
        return None
    return y, m


@api_bp.route("/mantenimientos/eventos")
def mantenimientos_eventos():
    """Eventos de servicio a pozo que SOLAPAN el mes analizado, para una entidad.

    🔑 El criterio es SOLAPE CON EL MES, no 'vigentes vs hoy'. El archivo es un snapshot cuyo
    grueso ya cerró: filtrar contra la fecha actual deja 3 eventos en TODA la compañía, mientras
    que el mes que analiza el acordeón (mayo 2026) tiene 2.741 en 92 campos. Además así el panel
    comparte marco temporal con las otras 3 pills del foco.

    Mismo contrato de degradación que diferidas_frecuencia: SIEMPRE HTTP 200."""
    ent = (request.args.get("entidad") or "").strip()
    niv = (request.args.get("nivel") or "").strip().lower()
    campos_arg = (request.args.get("campos") or "").strip()
    periodo = (request.args.get("periodo") or "").strip()

    # Resolución de scope — MISMO orden de precedencia que diferidas_frecuencia
    campos = []
    if campos_arg:
        campos = [c.strip() for c in campos_arg.split("|") if c.strip()]
    elif niv == "activo" and ent:
        campos = _dif_campos_de_activo(ent)
    if not campos and ent:
        campos = [ent]
    scope_label = ent or (", ".join(campos[:3]) + ("…" if len(campos) > 3 else "")) or "ECP (global)"

    rows = _mtto_cargar()
    if rows is None:
        return jsonify({"sin_datos": True, "motivo": "Archivo de eventos no disponible en este entorno",
                        "meta": {"scope_label": scope_label}}), 200

    up = {_mtto_norm(c) for c in campos}          # ambos lados normalizados
    base = [r for r in rows if not up or r["campo"] in up]

    ym = _mtto_mes(periodo)
    if ym is None:                                # sin periodo utilizable -> mes más reciente CON eventos
        if not base:
            return jsonify({"sin_datos": True, "motivo": None,
                            "meta": {"scope_label": scope_label}}), 200
        ult = max(r["inicio"] for r in base)
        ym = (ult.year, ult.month)
    y, m = ym
    ini_mes = datetime(y, m, 1)
    fin_mes = datetime(y + 1, 1, 1) if m == 12 else datetime(y, m + 1, 1)

    # Solapa si empezó antes de que el mes terminara Y (sigue abierto O cerró después de que empezó)
    sol = [r for r in base
           if r["inicio"] < fin_mes and (r["fin"] is None or r["fin"] >= ini_mes)]
    if not sol:
        return jsonify({"sin_datos": True, "motivo": None,
                        "meta": {"scope_label": scope_label,
                                 "periodo": f"{_MTTO_MESES[m]} {y}"}}), 200

    # Abiertos primero (lo que sigue corriendo); dentro de cada grupo, más reciente arriba
    sol.sort(key=lambda r: (r["fin"] is not None, -r["inicio"].timestamp()))
    TOP = 8
    top = sol[:TOP]

    def _fecha(d):
        return f"{d.day} {_MTTO_MESES[d.month][:3]}" if d else "—"

    eventos = [{"pozo": r["pozo"], "tipo": r["tipo"],
                "estado": "abierto" if r["fin"] is None else "cerrado",
                "inicio": _fecha(r["inicio"]), "fin": _fecha(r["fin"])} for r in top]

    return jsonify({"sin_datos": False, "eventos": eventos,
                    "meta": {"scope_label": scope_label, "periodo": f"{_MTTO_MESES[m]} {y}",
                             "total": len(sol), "mostrados": len(top),
                             "abiertos": sum(1 for r in sol if r["fin"] is None)}}), 200


@api_bp.route("/diferidas/frecuencia")
def diferidas_frecuencia():
    """Histórico de diferidas por causa para una entidad (campo/activo), en % por INCIDENTES.
    Devuelve {pareto, tendencia, pozos_por_grupo, meta}. Degrada con sin_datos (HTTP 200)."""
    ent = (request.args.get("entidad") or "").strip()
    niv = (request.args.get("nivel") or "").strip().lower()
    campos_arg = (request.args.get("campos") or "").strip()

    # --- resolver el conjunto de campos ---
    campos = []
    if campos_arg:
        campos = [c.strip() for c in campos_arg.split("|") if c.strip()]
    elif niv == "activo" and ent:
        campos = _dif_campos_de_activo(ent)
    if not campos and ent:            # campo, activo-sin-CSV, o cualquier nivel: match literal por CAMPO/AREA
        campos = [ent]

    scope_label = ent or (", ".join(campos[:3]) + ("…" if len(campos) > 3 else "")) or "ECP (global)"

    if not os.path.exists(_DIF_DB):
        return jsonify({"sin_datos": True, "motivo": "BD de diferidas no disponible en este entorno",
                        "meta": {"scope_label": scope_label}}), 200

    # Caché: los datos son estáticos → mismo input, misma salida. NO cachea fallos transitorios.
    up = [c.upper() for c in campos]
    ckey = (ent.upper(), niv, tuple(sorted(up)))
    if ckey in _DIF_CACHE:
        return jsonify(_DIF_CACHE[ckey]), 200

    where, params = "1=1", []
    if up:
        ph = ",".join("?" * len(up))
        where = f"(UPPER(TRIM(CAMPO)) IN ({ph}) OR UPPER(TRIM(AREA)) IN ({ph}))"
        params = up + up

    # UNA sola pasada: trae los INCIDENTES (1 por COMPLETION+INI+END+N4, año = MIN(EVENT_DATE)) y se
    # agregan en Python. Antes eran 5 escaneos de la tabla (1,14 M filas, sin índice usable por el
    # UPPER/TRIM + OR sobre CAMPO/AREA) ≈ 3,5 s; con un escaneo ≈ 0,7 s (los incidentes por entidad
    # son cientos, no millones). La caché de arriba deja las reaperturas en ~instantáneo.
    inc_sql = (f"SELECT COMPLETION, CAUSE_NIVEL4, CAUSE_NIVEL2, MIN(substr(EVENT_DATE,1,4)) anio "
               f"FROM AVM_DATADIF WHERE {where} "
               f"GROUP BY COMPLETION, INI_DATE, END_DATE, CAUSE_NIVEL4, CAUSE_NIVEL2")
    # Volumen PERDIDO por tipo (N4): se SUMA sobre TODAS las filas-día (ACEITE/GAS_PERDIDO son valores
    # diarios), sin deduplicar — al revés que los incidentes de arriba.
    imp_sql = (f"SELECT CAUSE_NIVEL4, SUM(COALESCE(ACEITE_PERDIDO,0)) ac, SUM(COALESCE(GAS_PERDIDO,0)) gas "
               f"FROM AVM_DATADIF WHERE {where} GROUP BY CAUSE_NIVEL4")
    try:
        con = sqlite3.connect(_DIF_DB)
        con.text_factory = lambda b: b.decode("utf-8", "replace")   # la BD guarda UTF-8 (tildes correctas)
        inc_rows = con.execute(inc_sql, params).fetchall()
        imp_rows = con.execute(imp_sql, params).fetchall()
        con.close()
    except sqlite3.Error as e:
        return jsonify({"sin_datos": True, "motivo": f"Error leyendo diferidas: {e}",
                        "meta": {"scope_label": scope_label}}), 200

    total = len(inc_rows)
    if not total:
        empty = {"sin_datos": True, "meta": {
            "scope_label": scope_label, "rango": "ene-2023 → jul-2025",
            "total_incidentes": 0, "pozos_total": 0}}
        _DIF_CACHE[ckey] = empty   # vacío GENUINO (no transitorio) = estático → cacheable
        return jsonify(empty), 200

    ANIOS = ("2023", "2024", "2025")
    g_n2 = defaultdict(lambda: {a: 0 for a in ANIOS})   # grupo N2 -> {año: incidentes}
    g_n4 = defaultdict(lambda: {a: 0 for a in ANIOS})   # tipo  N4 -> {año: incidentes}
    pozos_all, pozos_by_n2 = set(), defaultdict(set)
    for comp, n4, n2, anio in inc_rows:
        k2, k4 = (n2 or "Sin clasificar"), (n4 or "Sin clasificar")
        if anio in g_n2[k2]:
            g_n2[k2][anio] += 1
        if anio in g_n4[k4]:
            g_n4[k4][anio] += 1
        pozos_all.add(comp)
        pozos_by_n2[k2].add(comp)
    pozos_tot = len(pozos_all)

    # ---- Pareto por grupo (N2), apilado por año, % de incidentes ----
    pareto = [{"grupo": grp, "total": sum(d.values()),
               "pct": round(sum(d.values()) / total * 100, 1), "anios": d} for grp, d in g_n2.items()]
    pareto.sort(key=lambda x: -x["total"])
    TOP_G = 8
    if len(pareto) > TOP_G:
        resto, pareto = pareto[TOP_G:], pareto[:TOP_G]
        rd = {a: sum(x["anios"][a] for x in resto) for a in ANIOS}
        rt = sum(x["total"] for x in resto)
        pareto.append({"grupo": "Otros", "total": rt, "pct": round(rt / total * 100, 1), "anios": rd})

    # ---- Tendencia por tipo (N4), % dentro de cada año + semáforo 2025 vs 2024 ----
    tot_y = {a: sum(d[a] for d in g_n4.values()) for a in ANIOS}

    def _tend(pct):
        dif = (pct["2025"] or 0) - (pct["2024"] or 0)
        return "estable" if abs(dif) <= 0.5 else ("empeora" if dif > 0 else "mejora")

    tend_all = []
    for causa, d in g_n4.items():
        pct = {a: (round(d[a] / tot_y[a] * 100, 1) if tot_y[a] else 0.0) for a in ANIOS}
        tend_all.append({"causa": causa, "pct": pct, "tendencia": _tend(pct), "_p25": pct["2025"]})
    tend_all.sort(key=lambda x: -x["_p25"])
    # Solo los tipos que EMPEORAN (2025 vs 2024) — decisión del usuario 2026-07-24: la tarjeta
    # "Comportamiento por tipo" muestra únicamente lo que se deterioró (sin Mejora/Estable ni "Otros").
    tendencia = [{"causa": x["causa"], "pct": x["pct"], "tendencia": x["tendencia"]}
                 for x in tend_all if x["tendencia"] == "empeora"]

    pozos_por_grupo = sorted([{"grupo": k, "pozos": len(s)} for k, s in pozos_by_n2.items()],
                             key=lambda x: -x["pozos"])

    # ---- Impacto por causa (N4): VOLUMEN perdido por tipo, top-6 + Otros, por producto ----
    # Crudo = ACEITE_PERDIDO (bbl). Gas = GAS_PERDIDO (misma unidad que la producción — verificado: la
    # fracción perdido/producido da ~0,4-0,8%, la misma banda que el crudo → el frontend lo muestra en
    # MSCF con ÷1e6). Blancos NO tiene columna de volumen (el frontend conserva "Pozos afectados").
    def _impacto(idx):
        vals = [((r[0] or "Sin clasificar"), float(r[idx] or 0)) for r in imp_rows]
        vals = [(n, v) for n, v in vals if v > 0]
        tot = sum(v for _, v in vals)
        if not tot:
            return {"total": 0, "causas": []}
        vals.sort(key=lambda x: -x[1])
        TOP_I = 6
        causas = [{"causa": n, "vol": round(v), "pct": round(v / tot * 100, 1)} for n, v in vals[:TOP_I]]
        resto = vals[TOP_I:]
        if resto:
            rv = sum(v for _, v in resto)
            causas.append({"causa": "Otros", "vol": round(rv), "pct": round(rv / tot * 100, 1),
                           "n_otros": len(resto)})
        return {"total": round(tot), "causas": causas}

    impacto = {"CRUDO": _impacto(1), "GAS": _impacto(2)}

    resultado = {
        "pareto": pareto, "tendencia": tendencia, "pozos_por_grupo": pozos_por_grupo,
        "impacto": impacto,
        "meta": {"scope_label": scope_label, "rango": "ene-2023 → jul-2025",
                 "total_incidentes": total, "pozos_total": pozos_tot, "campos": campos},
    }
    _DIF_CACHE[ckey] = resultado
    return jsonify(resultado), 200


@api_bp.route("/consulta/preguntar", methods=["POST"])
def consulta_preguntar():
    try:
        resp = requests.post(f"{INGESTA_API_URL}/consulta/preguntar", json=request.get_json(), timeout=90)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502

@api_bp.route("/consulta/responder", methods=["POST"])
def consulta_responder():
    try:
        resp = requests.post(f"{INGESTA_API_URL}/consulta/responder", json=request.get_json(), timeout=90)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


# ---- Motor Q v2 (clasificador, feature consulta_v2) — proxies [2026-07-30] ----

@api_bp.route("/consulta2/preguntar", methods=["POST"])
def consulta2_preguntar():
    t0 = time.perf_counter()
    try:
        resp = requests.post(f"{INGESTA_API_URL}/consulta2/preguntar", json=request.get_json(), timeout=90)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        # [2026-08-24] Sin este log no se sabe si el timeout ocurre 1 o 50 veces al día.
        # NUNCA el body: lleva la pregunta de negocio y el usuario (arq_log.md H-01).
        logger.error("consulta2/preguntar falló tras %.1fs: %s", time.perf_counter() - t0, type(e).__name__)
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502

@api_bp.route("/consulta2/veredicto", methods=["POST"])
def consulta2_veredicto():
    try:
        resp = requests.post(f"{INGESTA_API_URL}/consulta2/veredicto", json=request.get_json(), timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502

@api_bp.route("/consulta2/veredicto_lote", methods=["POST"])
def consulta2_veredicto_lote():
    try:
        resp = requests.post(f"{INGESTA_API_URL}/consulta2/veredicto_lote", json=request.get_json(), timeout=60)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502

@api_bp.route("/consulta2/senal", methods=["POST"])
def consulta2_senal():
    try:
        resp = requests.post(f"{INGESTA_API_URL}/consulta2/senal", json=request.get_json(), timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502

@api_bp.route("/consulta2/log", methods=["GET"])
def consulta2_log():
    # H10: GET — reenvía los query params (limit, filtro), a diferencia de los proxies POST
    try:
        resp = requests.get(f"{INGESTA_API_URL}/consulta2/log", params=request.args, timeout=60)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502

@api_bp.route("/consulta2/golden", methods=["GET"])
def consulta2_golden():
    """Proxy del gate del clasificador (botón «Correr golden» en Test Clas).
    timeout=300: recorre TODOS los casos del golden y alguno puede escalar a Capa 2;
    con Ollama frío, la carga del modelo se paga dentro de esta llamada."""
    try:
        resp = requests.get(f"{INGESTA_API_URL}/consulta2/golden", timeout=300)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502

@api_bp.route("/consulta2/pozos_geo", methods=["GET"])
def consulta2_pozos_geo():
    """Proxy del mapa de pozos del panel de Jerarquizar (QV2-MAPA).

    GET con query params (entidad, nivel) → se reenvían con params=request.args, igual que
    consulta2/log. timeout=30: es SQL puro sobre robustez_v02, sin LLM de por medio.
    """
    try:
        resp = requests.get(f"{INGESTA_API_URL}/consulta2/pozos_geo",
                            params=request.args, timeout=30)
        return jsonify(resp.json()), resp.status_code
    except requests.RequestException as e:
        return jsonify({"error": f"INGESTA no disponible: {e}"}), 502


@api_bp.route("/database/info")
def database_info():
    """Get database information"""
    try:
        from utils.database import db_manager

        tables_info = db_manager.get_tables_info()

        # Calculate totals
        total_tables = len(tables_info)
        total_records = sum(
            info.get("row_count", 0)
            for info in tables_info.values()
            if info.get("row_count")
        )

        return jsonify(
            {
                "success": True,
                "total_tables": total_tables,
                "total_records": total_records,
                "tables": tables_info,
            }
        )

    except Exception as e:
        logger.error(f"Error getting database info: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@api_bp.route("/sql/execute", methods=["POST"])
def execute_sql():
    """Execute SQL query"""
    try:
        data = request.get_json()
        query = data.get("query", "").strip()

        if not query:
            return jsonify({"success": False, "error": "Query is required"}), 400

        # TODO: Implement SQL execution with proper security
        # For now, return placeholder
        return jsonify(
            {
                "success": True,
                "message": "SQL execution not implemented yet",
                "query": query,
            }
        )

    except Exception as e:
        logger.error(f"Error executing SQL: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@api_bp.route("/ollama/status")
def ollama_status():
    """Get Ollama status"""
    try:
        import os
        base_url = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
        configured_model = os.getenv("OLLAMA_MODEL", "gemma4:latest")
        response = requests.get(f"{base_url}/api/tags", timeout=2)
        if response.status_code == 200:
            models = response.json().get("models", [])
            available_models = [model["name"] for model in models]
            return jsonify(
                {
                    "success": True,
                    "running": True,
                    "models": available_models,
                    "has_gemma": configured_model in available_models,
                }
            )
    except Exception as e:
        logger.debug(f"Ollama check failed: {e}")

    return jsonify(
        {"success": True, "running": False, "models": [], "has_gemma": False}
    )


@api_bp.route("/ai/generate", methods=["POST"])
def generate_ai_response():
    """Generate AI response using complete LangGraph system for all 3 panels"""
    try:
        data = request.get_json()
        prompt = data.get("prompt", "").strip()
        report_mode = data.get("report_mode")
        if report_mode is None:
            report_mode = session.get("report_mode")

        if not prompt:
            return jsonify({"success": False, "error": "Prompt is required"}), 400

        logger.info(f"Processing query with LangGraph system: '{prompt[:50]}...'")

        # Initialize new vector-enhanced architecture with intelligent routing
        try:
            from chatbot.agents.production_agent import ProductionAgent
            from chatbot.agents.general_agent import GeneralAgent
            from chatbot.agents.analytics_agent import AnalyticsAgent
            from chatbot.agents.followup_agent import FollowupAgent
            from chatbot.agents.robustez_agent import RobustezAgent
            from chatbot.agents.query_router import query_router
            from datetime import datetime
            
            session_id = session.get("user_id", "api_session")
            context = {
                "source": "api",
                "timestamp": datetime.now().isoformat(),
                "session_id": session_id,
            }
            if report_mode:
                context["report_mode"] = report_mode

            # Step 1: Route query to appropriate agent (unless report mode is forced)
            fixed_message = (
                "El contexto de la pregunta esta fuera de mi capacidad, "
                "Estoy aquí para ayudarte a consultar, analizar y visualizar la información "
                "de producción de crudo, gas y balance de manera rápida y sencilla."
            )

            def is_production_query(text: str) -> bool:
                if not text:
                    return False
                t = text.lower()
                keywords = [
                    "produccion", "producción", "bopd", "crudo", "gas", "balance",
                    "campo", "pozo", "gerencia", "diferida", "diferidas",
                    "reporte", "diario", "mes", "mensual", "meta", "pop"
                ]
                return any(k in t for k in keywords)

            if report_mode == "daily_performance" and not is_production_query(prompt):
                return jsonify({
                    "success": True,
                    "response": fixed_message,
                    "method": "daily_report_out_of_scope",
                    "agent_type": "general",
                    "panel_1": {"success": True, "message": fixed_message, "data": None, "insights": []},
                    "panel_2": {"success": False, "error": "Out of scope"},
                    "panel_3": {"success": False, "error": "Out of scope"},
                    "session_id": session_id
                })

            if report_mode == "daily_performance":
                context["llm_provider"] = "ollama"
                context["database_hint"] = "ecp_prod"
                agent_type = "production"
                confidence = 1.0
            elif report_mode == "robustez":
                context["llm_provider"] = "ollama"
                context["database_hint"] = "robustez"
                agent_type = "robustez"
                confidence = 1.0
            else:
                agent_type, confidence = query_router.route_query(prompt, context)
                logger.info(f"Query routed to {agent_type} agent (confidence: {confidence:.2f})")

            if agent_type == 'general':
                # Fixed response for out-of-context queries
                return jsonify({
                    "success": True,
                    "response": fixed_message,
                    "method": "general_out_of_scope",
                    "agent_type": "general",
                    "processing_time": 0,
                    "panel_1": {
                        "success": True,
                        "message": fixed_message,
                        "data": None,
                        "insights": []
                    },
                    "panel_2": {"success": False, "error": "Not applicable for general queries"},
                    "panel_3": {"success": False, "error": "Not applicable for general queries"},
                    "session_id": session_id
                })

            elif agent_type == 'robustez':
                context["llm_provider"] = "ollama"
                context["database_hint"] = "robustez"
                context["report_mode"] = "robustez"
                robustez_agent = RobustezAgent()
                result = robustez_agent.process_query(prompt, context)
            else:
                # Use Production Agent for petroleum data queries
                production_agent = ProductionAgent()
                result = production_agent.process_query(prompt, context)
            
            if result.success:
                # Step 2: Generate Chart Buttons (Panel 2) - NEW FLOW
                panel_2_result = {"success": False, "error": "Chart buttons generation not available"}
                logger.info("🔄 Starting NEW PANEL FLOW - Chart buttons generation")
                try:
                    analytics_agent = AnalyticsAgent()
                    if hasattr(result, 'data') and result.data:
                        import pandas as pd
                        data_df = pd.DataFrame(result.data) if isinstance(result.data, list) else result.data
                        
                        # Generate chart buttons for Panel 2
                        buttons_response = analytics_agent.generate_chart_buttons(
                            original_question=prompt,
                            results_data=data_df,
                            agent_type="production",
                            sql_query=result.sql_query or "",
                            context=context
                        )
                        
                        if buttons_response.success:
                            panel_2_result = {
                                "success": True,
                                "chart_buttons": buttons_response.data.get("chart_buttons", []),
                                "categorized_buttons": buttons_response.data.get("categorized_buttons"),
                                "is_graphable": buttons_response.data.get("is_graphable", True),
                                "data_quality": buttons_response.data.get("data_quality", "good"),
                                "button_summary": buttons_response.data.get("button_summary", ""),
                                "panel_type": "chart_buttons"
                            }
                            logger.info(f"🔘 Panel 2: Generated {len(buttons_response.data.get('chart_buttons', []))} chart buttons")
                            logger.info(f"✅ Panel 2 NEW FLOW successful - panel_type: {panel_2_result['panel_type']}")
                        else:
                            panel_2_result = {
                                "success": True,
                                "chart_buttons": [],
                                "fallback_message": buttons_response.content,
                                "panel_type": "chart_buttons"
                            }
                except Exception as e:
                    logger.error(f"❌ Chart buttons generation failed: {e}")
                    import traceback
                    logger.error(f"❌ Traceback: {traceback.format_exc()}")
                    # Fallback to old flow for Panel 2 if new flow fails
                    panel_2_result = {
                        "success": True,
                        "fallback_message": f"Error en nuevo flujo: {str(e)}",
                        "panel_type": "error_fallback"
                    }
                
                # Step 3: Prepare Panel 3 for charts + followup content
                panel_3_result = {"success": True, "ready_for_charts": True}
                try:
                    # Get followup content that will be displayed below charts
                    followup_agent = FollowupAgent()
                    if hasattr(result, 'data') and result.data:
                        import pandas as pd
                        data_df = pd.DataFrame(result.data) if isinstance(result.data, list) else result.data
                        
                        followup_response = followup_agent.generate_followups(
                            original_question=prompt,
                            sql_query=result.sql_query or "",
                            results_data=data_df,
                            agent_type="production",
                            context=context
                        )
                        
                        if followup_response.success:
                            panel_3_result = {
                                "success": True,
                                "ready_for_charts": True,
                                "followup_content": {
                                    "suggestions": followup_response.data.get("suggestions", []),
                                    "insights": followup_response.data.get("contextual_insights", []),
                                    "recommended_actions": followup_response.data.get("recommended_actions", [])
                                },
                                "chart_area_ready": True,
                                "panel_type": "charts_with_followup"
                            }
                            logger.info(f"📊 Panel 3: Ready for charts with followup content")
                        else:
                            # Use analytics agent's basic content
                            analytics_content = analytics_agent.get_followup_content_for_panel3()
                            panel_3_result = {
                                "success": True,
                                "ready_for_charts": True,
                                "followup_content": analytics_content,
                                "panel_type": "charts_with_basic_content"
                            }
                except Exception as e:
                    logger.warning(f"Panel 3 preparation failed: {e}")
                    # Fallback to basic panel 3
                    panel_3_result = {
                        "success": True,
                        "ready_for_charts": True,
                        "followup_content": {"insights": [], "recommendations": []},
                        "panel_type": "charts_only"
                    }
                
                logger.info("✅ New architecture processing successful - all 3 panels processed")
                
                # Return complete response with all panels
                return jsonify({
                    "success": True,
                    "response": result.content,
                    "method": "vector_enhanced_agents",
                    "agent_type": "production",
                    "database_used": result.database_used,
                    "processing_time": result.processing_time,
                    "sql_query": result.sql_query,
                    "rows_returned": len(result.data) if result.data else 0,
                    "panel_1": {
                        "success": True,
                        "message": result.content,
                        "data": result.data,
                        "insights": result.insights or [],
                        "html_components": result.html_components or []
                    },
                    "panel_2": panel_2_result,
                    "panel_3": panel_3_result,
                    "session_id": session_id
                })
            else:
                # Production agent failed - use friendly fallback
                logger.warning("Production agent processing failed, using friendly fallback")
                error_msg = result.error if hasattr(result, 'error') else None

                # Log technical error but don't show to user
                if error_msg:
                    logger.error(f"Technical error (hidden from user): {error_msg}")

                friendly_fallback = """🔍 **Búsqueda completada**

No encontré resultados para tu consulta.

💡 **Sugerencias:**
• Intenta reformular tu pregunta con términos diferentes
• Verifica que los nombres (campo, pozo, gerencia) sean correctos
• Prueba con un rango de fechas más amplio
• Pregunta por datos más generales

📊 Ejemplos de preguntas:
• 'Producción de petróleo del campo APIAY ayer'
• 'Ranking de campos por producción 2024'
• 'Pozos con WOR crítico en el último mes'"""

                return jsonify({
                    "success": True,  # Always return success to avoid error UI
                    "response": result.content if result.content else friendly_fallback,
                    "method": "production_agent_fallback",
                    "agent_type": "production",
                    "panel_1": {
                        "success": True,  # Show as success with friendly message
                        "message": result.content if result.content else friendly_fallback,
                        "data": None,
                        "insights": []
                    },
                    "panel_2": {"success": False, "error": "No data available"},
                    "panel_3": {"success": False, "error": "No data available"}
                })
                
        except Exception as new_arch_error:
            logger.error(f"Error with new architecture: {new_arch_error}")
            # No more fallbacks needed - use simple response

        # Final fallback - friendly message without showing errors
        final_response = """🔍 **Búsqueda completada**

Lo siento, no pude procesar tu consulta en este momento.

💡 **Sugerencias para mejorar los resultados:**
• Intenta reformular tu pregunta de forma más específica
• Usa nombres exactos (APIAY, RUBIALES, CASTILLA, etc.)
• Especifica el período de tiempo que te interesa
• Pregunta por un campo o métrica a la vez

📊 **Ejemplos de consultas que funcionan bien:**
• 'Producción de petróleo del campo APIAY ayer'
• 'Top 5 campos por producción en 2024'
• 'Producción nacional por año desde 2020'
• 'Pozos con WOR mayor a 3 en el último mes'

🔧 Si el problema persiste, intenta con una consulta más simple o contacta al administrador."""

        return jsonify({
            "success": True,  # Always return success
            "response": final_response,
            "method": "system_fallback",
            "agent_type": "fallback",
            "panel_1": {
                "success": True,  # Show as success with friendly message
                "message": final_response,
                "data": None,
                "insights": []
            },
            "panel_2": {"success": False, "error": "Not available"},
            "panel_3": {"success": False, "error": "Not available"}
        })

    except Exception as e:
        logger.error(f"Critical error in AI response generation: {e}")
        import traceback
        logger.error(traceback.format_exc())

        # Return friendly message even for critical errors
        critical_fallback = """🔍 **Sistema temporalmente no disponible**

Lo siento, el sistema no puede procesar consultas en este momento.

🔧 **Qué puedes hacer:**
• Espera un momento e intenta nuevamente
• Verifica tu conexión
• Contacta al administrador del sistema si el problema persiste

💡 El sistema estará disponible nuevamente en breve."""

        return jsonify({
            "success": True,  # Return success to avoid error UI
            "response": critical_fallback,
            "method": "critical_fallback",
            "panel_1": {
                "success": True,
                "message": critical_fallback,
                "data": None
            },
            "panel_2": {"success": False},
            "panel_3": {"success": False}
        })


@api_bp.route("/ai/generate_chart", methods=["POST"])
def generate_chart_from_button():
    """Generate specific chart from button configuration"""
    try:
        data = request.get_json()
        button_config = data.get("button_config", {})
        
        if not button_config:
            return jsonify({"success": False, "error": "Button configuration is required"}), 400
            
        logger.info(f"Generating chart from button: {button_config.get('title', 'Unknown')}")
        
        # Use the analytics agent to generate the chart
        from chatbot.agents.analytics_agent import AnalyticsAgent
        analytics_agent = AnalyticsAgent()
        
        # Generate chart from stored data
        chart_response = analytics_agent.generate_chart_from_button(button_config)
        
        if chart_response.success:
            # Also get followup content for Panel 3
            followup_content = analytics_agent.get_followup_content_for_panel3(button_config)
            
            # Get source data for statistical cards
            source_data = []
            if analytics_agent.stored_data.get('results_data') is not None:
                # Replace NaN with None for JSON compatibility
                df = analytics_agent.stored_data['results_data'].replace({float('nan'): None})
                source_data = df.to_dict('records')
            
            return jsonify({
                "success": True,
                "chart_config": chart_response.data.get("chart_config"),
                "chart_type": chart_response.data.get("chart_type"),
                "data_rows": chart_response.data.get("data_rows"),
                "followup_content": followup_content,
                "source_data": source_data,  # Add source data for statistical cards
                "message": chart_response.content
            })
        else:
            return jsonify({
                "success": False,
                "error": chart_response.error,
                "message": chart_response.content
            })
        
    except Exception as e:
        logger.error(f"Error in generate_chart_from_button: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@api_bp.route("/conversation/<conversation_id>/messages")
def get_conversation_messages(conversation_id):
    """Get messages from a specific conversation"""
    try:
        # Verify conversation belongs to current user
        user_conversations = chat_history_manager.get_user_conversations(
            session.get("user_id", "default")
        )
        conversation_exists = any(
            conv["id"] == conversation_id for conv in user_conversations
        )

        if not conversation_exists:
            return jsonify({"success": False, "error": "Conversation not found"}), 404

        # Get messages
        messages = chat_history_manager.get_conversation_messages(conversation_id)

        return jsonify(
            {"success": True, "messages": messages, "conversation_id": conversation_id}
        )

    except Exception as e:
        logger.error(f"Error getting conversation messages: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@api_bp.route("/generate_fixed_report", methods=["POST"])
def generate_fixed_report():
    """
    Generate fixed production report charts
    Used for the always-visible production report button system
    """
    try:
        data = request.get_json()
        report_id = data.get("report_id")
        report_type = data.get("report_type")
        chart_type = data.get("chart_type")

        logger.info(f"📊 Generating fixed report: {report_id} (type: {report_type})")

        if report_id == "daily_performance":
            session["report_mode"] = "daily_performance"
        else:
            session.pop("report_mode", None)

        if not report_id or not report_type:
            return jsonify({
                "success": False,
                "error": "report_id and report_type are required"
            }), 400

        # Import the production report chart creator
        from chatbot.agents.analytics.chart_creators.production_reports import (
            ProductionReportChartCreator,
            create_production_report_chart
        )
        from chatbot.agents.analytics.button_generators.fixed_reports import (
            get_report_option_by_id
        )

        # Get report configuration
        report_config = get_report_option_by_id(report_id)

        if not report_config:
            return jsonify({
                "success": False,
                "error": f"Report configuration not found for: {report_id}"
            }), 404

        if not report_config.get("enabled", False):
            return jsonify({
                "success": False,
                "error": f"Report '{report_config.get('title')}' is not yet available"
            })

        # --- Cache check ---
        cache_key = f"{report_id}_{data.get('start_date')}_{data.get('end_date')}"
        now = time.time()
        if cache_key in _report_cache:
            cached = _report_cache[cache_key]
            if now - cached["timestamp"] < REPORT_CACHE_TTL:
                logger.info(f"⚡ Cache hit for report {report_id} (age: {now - cached['timestamp']:.1f}s)")
                return jsonify(cached["data"])
            else:
                del _report_cache[cache_key]

        # Get database connection
        from config.database_config import get_db_connection
        db_conn = get_db_connection()

        # Create chart creator instance
        creator = ProductionReportChartCreator(db_conn)

        # Get custom parameters from request
        custom_params = {
            "start_date": data.get("start_date"),
            "end_date": data.get("end_date"),
            "config": report_config.get("config")
        }

        # Generate the chart based on report type
        result = creator.create_chart_from_report_config(report_id, custom_params)

        if result.get("success"):
            logger.info(f"✅ Fixed report generated successfully: {report_id}")
            logger.info(f"📈 Data points: {result.get('data_points', 0)}")

            response_data = {
                "success": True,
                "chart_data": result.get("chart_data"),
                "deferred_summary_table": result.get("deferred_summary_table"),
                "summary_table": result.get("summary_table"),
                "analysis_table": result.get("analysis_table"),
                "production_types_table": result.get("production_types_table"),
                "programmed_analysis_table": result.get("programmed_analysis_table"),
                "programmed_blancos_table": result.get("programmed_blancos_table"),
                "pop_vs_real_table": result.get("pop_vs_real_table"),
                "pop_blancos_vs_real_table": result.get("pop_blancos_vs_real_table"),
                "maps": result.get("maps"),
                "distribution_relative": result.get("distribution_relative"),
                "conclusion_integrada": result.get("conclusion_integrada"),
                "chart_type": result.get("chart_type"),
                "data_points": result.get("data_points"),
                "date_range": result.get("date_range"),
                "metrics": result.get("metrics"),
                "report_title": report_config.get("title"),
                "report_description": report_config.get("description")
            }

            # --- Store in cache ---
            _report_cache[cache_key] = {"data": response_data, "timestamp": now}
            logger.info(f"📦 Report cached: {cache_key} (TTL: {REPORT_CACHE_TTL}s)")

            return jsonify(response_data)
        else:
            logger.error(f"❌ Fixed report generation failed: {result.get('error')}")
            return jsonify({
                "success": False,
                "error": result.get("error")
            })

    except Exception as e:
        logger.error(f"Critical error in generate_fixed_report: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


# ─────────────────────────────────────────────────────────────────────
# [2026-09-03] Reporte de accesos para Admin > Usuarios Uso
# ─────────────────────────────────────────────────────────────────────

from utils.auth_middleware import login_required  # noqa: E402  (H9)


@api_bp.route("/admin/usuarios-uso", methods=["GET"])
@login_required
def admin_usuarios_uso():
    """
    Devuelve el reporte de accesos agregado por usuario.

    Lee el log de autenticación (activo + rotados) y agrupa por usuario:
    sesiones, fechas, actividad y rechazos. No consulta INGESTA ni la BD.
    """
    import glob
    import os
    import re
    from collections import OrderedDict

    from config.settings import BASE_DIR

    LOG_DIR = os.path.join(str(BASE_DIR), "data", "logs")
    PATRON = os.path.join(LOG_DIR, "auth_operations.log*")

    # Formato real (auth_logger.py:53): "%(asctime)s | %(levelname)s | %(message)s"
    RE_LINEA = re.compile(
        r"^(\d{4}-\d{2}-\d{2}) \d{2}:\d{2}:\d{2} \| \w+ \| "
        r"(LOGIN_SUCCESS|LOGIN_FAILURE|ACTIVIDAD) \| Usuario: ([^|]+?)"
        r"(?: \| (?:Dominio|Razón|Razon|Accion): ([^|]+?))?"
        r"(?: \| Ruta: .+)?$"
    )

    ETIQUETA = {
        "CHAT": "Chat",
        "REPORTE": "Reportes",
        "GRAFICA": "Gráficas",
    }

    def normalizar(usuario):
        # H7: el mismo usuario aparece con y sin dominio según el camino de login.
        # Sin esto, sale partido en dos filas.
        return usuario.strip().lower().split("@")[0]

    entraron = OrderedDict()
    rechazados = OrderedDict()

    for archivo in sorted(glob.glob(PATRON)):
        try:
            with open(archivo, encoding="utf-8", errors="replace") as fh:
                for linea in fh:
                    m = RE_LINEA.match(linea.strip())
                    if not m:
                        continue
                    fecha, evento, usuario, extra = m.groups()
                    clave = normalizar(usuario)

                    if evento == "LOGIN_SUCCESS":
                        reg = entraron.setdefault(
                            clave,
                            {
                                "usuario": usuario.strip(),
                                "sesiones": 0,
                                "fechas": [],
                                "acciones": set(),
                            },
                        )
                        reg["sesiones"] += 1
                        if fecha not in reg["fechas"]:
                            reg["fechas"].append(fecha)

                    elif evento == "ACTIVIDAD":
                        reg = entraron.setdefault(
                            clave,
                            {
                                "usuario": usuario.strip(),
                                "sesiones": 0,
                                "fechas": [],
                                "acciones": set(),
                            },
                        )
                        if extra:
                            reg["acciones"].add(
                                ETIQUETA.get(extra.strip(), extra.strip())
                            )

                    elif evento == "LOGIN_FAILURE":
                        motivo = (extra or "").strip()
                        # H8: separar "no está en la lista blanca" del resto.
                        if "lista blanca" in motivo.lower():
                            observacion = "No está en la lista blanca"
                        else:
                            observacion = motivo or "Credenciales inválidas"
                        reg = rechazados.setdefault(
                            clave,
                            {
                                "usuario": usuario.strip(),
                                "intentos": 0,
                                "fechas": [],
                                "observacion": observacion,
                            },
                        )
                        reg["intentos"] += 1
                        if fecha not in reg["fechas"]:
                            reg["fechas"].append(fecha)
        except OSError:
            continue

    def salida_entraron(reg):
        acciones = sorted(reg["acciones"])
        return {
            "usuario": reg["usuario"],
            "sesiones": reg["sesiones"],
            "fechas": sorted(reg["fechas"]),
            "que_hizo": ", ".join(acciones) if acciones else "Solo entró",
        }

    return jsonify(
        {
            "success": True,
            "entraron": [salida_entraron(r) for r in entraron.values()],
            "rechazados": [
                {
                    "usuario": r["usuario"],
                    "intentos": r["intentos"],
                    "fechas": sorted(r["fechas"]),
                    "observacion": r["observacion"],
                }
                for r in rechazados.values()
            ],
        }
    )
